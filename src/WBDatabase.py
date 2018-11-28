from openpyxl import *

class Database():
    def __init__(self):
        try:
            self.wb = load_workbook('./data/excel.xlsx')
        except:
            self.wb = Workbook()
            sheet = self.wb.active
            # resize the width of columns in 
            # excel spreadsheet
            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 40
            sheet.column_dimensions['H'].width = 50
            
            # write given data to an excel spreadsheet 
            # at particular location 
            sheet.cell(row=1, column=1).value = "ID"
            sheet.cell(row=1, column=2).value = "Name"
            sheet.cell(row=1, column=3).value = "Course"
            sheet.cell(row=1, column=4).value = "Semester"
            sheet.cell(row=1, column=5).value = "Form Number"
            sheet.cell(row=1, column=6).value = "Contact Number"
            sheet.cell(row=1, column=7).value = "Email id"
            sheet.cell(row=1, column=8).value = "Address"
            self.wb.save('./data/excel.xlsx')
            
    def insert_data(self, name, course, sem, form_no, contact_no, email_id, address):
        sheet = self.wb.active
        current_row = sheet.max_row
        sheet.cell(row=current_row + 1, column=1).value = current_row
        sheet.cell(row=current_row + 1, column=2).value = name
        sheet.cell(row=current_row + 1, column=3).value = course 
        sheet.cell(row=current_row + 1, column=4).value = sem
        sheet.cell(row=current_row + 1, column=5).value = form_no 
        sheet.cell(row=current_row + 1, column=6).value = contact_no 
        sheet.cell(row=current_row + 1, column=7).value = email_id
        sheet.cell(row=current_row + 1, column=8).value = address
        self.wb.save('./data/excel.xlsx')
        print(current_row)
        return current_row

    def retrieve_names(self):
        sheet = self.wb.active
        name_array = []
        max_row = sheet.max_row
        for i in range(2, max_row+1):
            cell = "B" + str(i)
            name_array.append(sheet[cell].value)

        return name_array
